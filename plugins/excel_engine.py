"""
plugins/excel_engine.py
-----------------------
Professional Excel workbook (.xlsx) generator and analyzer using openpyxl.
Supports styled headers, formulas (SUM, AVERAGE, etc.), currency formatting, and auto column sizing.
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

PLUGIN = {
    "name": "excel_engine",
    "description": (
        "Creates or edits professional Excel spreadsheets (.xlsx) with styled headers, "
        "calculated formulas (SUM, AVERAGE, etc.), auto-sized columns, and multiple sheets. "
        "Use when user wants to create budgets, financial models, gradebooks, data tables, or analytics sheets."
    ),
    "parameters": {
        "type": "OBJECT",
        "properties": {
            "output_path": {
                "type": "STRING",
                "description": "Path to save the .xlsx file e.g. Desktop/budget_2026.xlsx",
            },
            "sheet_name": {
                "type": "STRING",
                "description": "Name of the sheet (default: 'Sheet1')",
            },
            "headers": {
                "type": "ARRAY",
                "description": "List of header column names e.g. ['Item', 'Category', 'Amount', 'Date']",
                "items": {"type": "STRING"},
            },
            "rows": {
                "type": "ARRAY",
                "description": "2D list of row values",
                "items": {
                    "type": "ARRAY",
                    "items": {"type": "STRING"},
                },
            },
            "add_totals_row": {
                "type": "BOOLEAN",
                "description": "Whether to auto-append a formula-driven Total row at the bottom",
            },
            "theme_color": {
                "type": "STRING",
                "description": "Header background hex color (default: '1F4E79')",
            },
        },
        "required": ["output_path", "headers", "rows"],
    },
}


def run(parameters: dict, player=None, speak=None) -> str:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return "❌ openpyxl is required for Excel operations. Please install with: pip install openpyxl"

    out_raw = parameters.get("output_path", "Desktop/spreadsheet.xlsx")
    out_p = Path(out_raw)
    if not out_p.is_absolute():
        out_p = Path.home() / "Desktop" / out_p.name

    out_p.parent.mkdir(parents=True, exist_ok=True)
    sheet_name = parameters.get("sheet_name", "Sheet1")
    headers = parameters.get("headers", [])
    rows = parameters.get("rows", [])
    add_totals = bool(parameters.get("add_totals_row", False))
    theme_hex = (parameters.get("theme_color") or "1F4E79").lstrip("#")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Styles
    header_fill = PatternFill(start_color=theme_hex, end_color=theme_hex, fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="D3D3D3"),
        right=Side(style="thin", color="D3D3D3"),
        top=Side(style="thin", color="D3D3D3"),
        bottom=Side(style="thin", color="D3D3D3"),
    )

    # 1. Write Headers
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=str(h))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[1].height = 26

    # 2. Write Data Rows
    current_row = 2
    for r_data in rows:
        ws.row_dimensions[current_row].height = 20
        for col_idx, val in enumerate(r_data, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            # Numeric conversion if feasible
            if isinstance(val, (int, float)):
                cell.value = val
            elif isinstance(val, str) and val.replace(".", "", 1).replace("-", "", 1).isdigit():
                cell.value = float(val) if "." in val else int(val)
            else:
                cell.value = val
            cell.font = Font(name="Calibri", size=10)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
        current_row += 1

    # 3. Optional Totals Row
    if add_totals and len(rows) > 0:
        ws.row_dimensions[current_row].height = 22
        total_font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=current_row, column=1, value="Total").font = total_font
        for c_idx in range(2, len(headers) + 1):
            col_letter = get_column_letter(c_idx)
            # check if first row in this col was numeric
            first_val = rows[0][c_idx - 1] if len(rows[0]) >= c_idx else None
            is_num = isinstance(first_val, (int, float)) or (isinstance(first_val, str) and first_val.replace(".", "", 1).replace("-", "", 1).isdigit())
            if is_num:
                formula = f"=SUM({col_letter}2:{col_letter}{current_row - 1})"
                cell = ws.cell(row=current_row, column=c_idx, value=formula)
                cell.font = total_font
                cell.border = thin_border

    # 4. Auto-fit Column Widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(str(out_p))
    return f"✅ Excel workbook successfully created with {len(rows)} rows: {out_p.name} ({out_p.resolve()})"
